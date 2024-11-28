import argparse
import openpyxl
from veracode_api_py.identity import Roles

class Role():
    name: str = ""
    description: str = ""
    is_api: bool = False
    jit_assignable: bool = False
    jit_assignable_default: bool = False
    roles = []
    permissions = []
    permission_types = []

    def __init__(self, name, description, user_type, jit_assignable, jit_assignable_default, role, permission, permission_types):
        self.name = name
        self.description = description
        self.is_api = user_type and user_type.upper().strip() == "API"
        self.jit_assignable = jit_assignable
        self.jit_assignable_default = jit_assignable_default
        self.roles = [{"role_description": role}] if role else []
        self.permissions = [{"permission_name": permission}] if permission else []
        if permission_types:
            self.permissions[0]["permission_types"] = [permission_type.strip() for permission_type in permission_types.split(",")]

def parse_line(excel_sheet, row) -> Role:
    return Role(
        name=excel_sheet.cell(row = row, column = 1).value,
        description=excel_sheet.cell(row = row, column = 2).value,
        user_type=excel_sheet.cell(row = row, column = 3).value,
        jit_assignable=excel_sheet.cell(row = row, column = 4).value,
        jit_assignable_default=excel_sheet.cell(row = row, column = 5).value,
        role=excel_sheet.cell(row = row, column = 6).value,
        permission=excel_sheet.cell(row = row, column = 7).value,
        permission_types=excel_sheet.cell(row = row, column = 8).value
    )

def parse_roles(file_name):
    excel_file = openpyxl.load_workbook(file_name)
    excel_sheet = excel_file.active
    roles = []
    current_role: Role = None
    try:
        for row in range(2, excel_sheet.max_row+1):
            current_line = parse_line(excel_sheet, row)
            if current_line.name:
                if current_role:
                    roles.append(current_role)
                current_role = current_line
            elif current_role:
                if current_line.permissions:
                    current_role.permissions.append(current_line.permissions[0])
                if current_line.roles:
                    current_role.roles.append(current_line.roles[0])   
        if current_role:
            roles.append(current_role)
        return roles
    finally:
        excel_file.close()

def main():
    parser = argparse.ArgumentParser(
        description="This script allows for the creation of Custom User Roles in Veracode."
    )
    parser.add_argument(
        "-f",
        "--file",
        help="File containing the information on the roles(s) to create.",
        required=True,
    )

    args = parser.parse_args()

    roles = parse_roles(args.file)

    for role in roles:
        print(f"Trying to create role {role.name}")
        try:
            Roles().create(
                role_name= role.name,
                role_description= role.description,
                is_api=role.is_api,
                jit_assignable=role.jit_assignable,
                jit_assignable_default=role.jit_assignable_default,
                permissions=role.permissions,
                child_roles=role.roles
            )
            print(f"Role {role.name} created successfully")
        except Exception as e:
            print(f"ERROR when creating role {role.name}")
        finally:
            print("--------------------------------------")


if __name__ == '__main__':
    main()
    